#!/usr/bin/env python3
"""
AWS Resource Excel Reporter
Scans all AWS resources across all regions and generates a detailed Excel report with:
- Resource inventory across all services
- Cost indicator (Yes/No)
- Creation date, last used date
- Delete recommendation for unused resources
- Full S3 bucket details (object count, size, last modified)
"""

import boto3
import argparse
import sys
import os
from datetime import datetime, timezone, timedelta
from botocore.exceptions import ClientError
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
except ImportError:
    print("❌ openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────
def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', start_color='1F3864')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            bottom=Side(style='thin', color='4472C4'),
            right=Side(style='thin', color='4472C4')
        )

def style_row(ws, row, cols, even=True):
    color = 'F2F7FF' if even else 'FFFFFF'
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name='Arial', size=9)
        cell.fill = PatternFill('solid', start_color=color)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = Border(
            bottom=Side(style='thin', color='D9E1F2'),
            right=Side(style='thin', color='D9E1F2')
        )

def style_title(ws, title, subtitle=""):
    ws.merge_cells('A1:P1')
    t = ws['A1']
    t.value = title
    t.font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
    t.fill = PatternFill('solid', start_color='1F3864')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    if subtitle:
        ws.merge_cells('A2:P2')
        s = ws['A2']
        s.value = subtitle
        s.font = Font(name='Arial', size=10, color='595959', italic=True)
        s.fill = PatternFill('solid', start_color='D9E1F2')
        s.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 20

def cost_color(ws, cell, value):
    if value == 'YES':
        cell.font = Font(name='Arial', bold=True, size=9, color='C00000')
        cell.fill = PatternFill('solid', start_color='FFE7E7')
    elif value == 'NO':
        cell.font = Font(name='Arial', bold=True, size=9, color='375623')
        cell.fill = PatternFill('solid', start_color='E7F3E7')

def recommend_color(cell, value):
    if value and 'DELETE' in value.upper():
        cell.font = Font(name='Arial', bold=True, size=9, color='C00000')
        cell.fill = PatternFill('solid', start_color='FFE7E7')
    elif value and 'KEEP' in value.upper():
        cell.font = Font(name='Arial', bold=True, size=9, color='375623')
        cell.fill = PatternFill('solid', start_color='E7F3E7')
    elif value and 'REVIEW' in value.upper():
        cell.font = Font(name='Arial', bold=True, size=9, color='7F4F00')
        cell.fill = PatternFill('solid', start_color='FFF2CC')


# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────
def get_name_tag(tags):
    if not tags:
        return None
    for t in tags:
        if t.get('Key') == 'Name':
            return t.get('Value')
    return None

def fmt_date(dt):
    if not dt:
        return 'N/A'
    if isinstance(dt, str):
        return dt[:10]
    try:
        # Strip timezone info — Excel does not support tz-aware datetimes
        if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)
        return dt.strftime('%Y-%m-%d')
    except:
        return str(dt)[:10]

def days_ago(dt):
    if not dt:
        return None
    try:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - dt).days
    except:
        return None

def fmt_bytes(size):
    if size is None:
        return 'N/A'
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size < 1024:
            return f"{size:.2f} {unit}"
        size /= 1024
    return f"{size:.2f} PB"

def safe(fn, default='N/A'):
    try:
        result = fn()
        return result if result is not None else default
    except:
        return default


# ─────────────────────────────────────────────────────────
# SCANNER
# ─────────────────────────────────────────────────────────
class AWSExcelReporter:
    def __init__(self, access_key, secret_key, session_token=None):
        self.session = boto3.Session(
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
            aws_session_token=session_token,
            region_name='us-east-1'
        )
        self.account_id = 'unknown'

    def client(self, service, region='us-east-1'):
        return self.session.client(service, region_name=region)

    def get_regions(self):
        try:
            ec2 = self.client('ec2')
            return [r['RegionName'] for r in ec2.describe_regions()['Regions']]
        except:
            return ['us-east-1', 'ap-south-1', 'eu-west-1', 'us-west-2']

    def is_active_region(self, region):
        try:
            ec2 = self.client('ec2', region)
            vpcs = ec2.describe_vpcs()['Vpcs']
            insts = ec2.describe_instances()['Reservations']
            return len(vpcs) > 1 or len(insts) > 0
        except:
            return False

    # ── EC2 INSTANCES ────────────────────────
    def get_ec2(self, region):
        rows = []
        try:
            ec2 = self.client('ec2', region)
            reservations = ec2.describe_instances()['Reservations']
            for r in reservations:
                for i in r['Instances']:
                    state = i.get('State', {}).get('Name', '')
                    if state == 'terminated':
                        continue
                    name = get_name_tag(i.get('Tags')) or i['InstanceId']
                    launch = i.get('LaunchTime')
                    d = days_ago(launch)
                    stopped = state == 'stopped'
                    stopped_days = None
                    if stopped:
                        try:
                            reason = i.get('StateTransitionReason', '')
                            if 'User initiated (' in reason:
                                date_str = reason.split('(')[1].split(')')[0]
                                stop_dt = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S GMT')
                                stopped_days = days_ago(stop_dt.replace(tzinfo=timezone.utc))
                        except:
                            pass

                    if stopped and stopped_days and stopped_days > 30:
                        rec = f'⚠️ DELETE — Stopped {stopped_days} days'
                    elif stopped:
                        rec = '🔍 REVIEW — Instance stopped'
                    elif state == 'running':
                        rec = '✅ KEEP — Running'
                    else:
                        rec = '🔍 REVIEW — Check state'

                    rows.append({
                        'Resource Type': 'EC2 Instance',
                        'Name': name,
                        'Resource ID': i['InstanceId'],
                        'Region': region,
                        'Status/State': state,
                        'Has Cost?': 'YES' if state == 'running' else 'NO (stopped)',
                        'Instance Type': i.get('InstanceType', 'N/A'),
                        'Created / Launched': fmt_date(launch),
                        'Last Used': fmt_date(i.get('StateTransitionReason', launch)),
                        'Days Old': d,
                        'Key Details': f"IP: {i.get('PrivateIpAddress','N/A')} | AZ: {i.get('Placement',{}).get('AvailabilityZone','N/A')}",
                        'Recommendation': rec,
                        'Notes': f"AMI: {i.get('ImageId','N/A')}"
                    })
        except ClientError:
            pass
        return rows

    # ── S3 BUCKETS ───────────────────────────
    def get_s3(self):
        rows = []
        try:
            s3 = self.client('s3')
            buckets = s3.list_buckets()['Buckets']
            print(f"   📦 Found {len(buckets)} S3 buckets — fetching details...")

            for b in buckets:
                name = b['Name']
                created = b.get('CreationDate')

                # Get bucket location
                region = safe(lambda: s3.get_bucket_location(
                    Bucket=name)['LocationConstraint'] or 'us-east-1', 'us-east-1')

                # Get object count and size via CloudWatch metrics
                obj_count = 'N/A'
                total_size = 'N/A'
                last_modified = 'N/A'

                try:
                    cw = self.client('cloudwatch', region if region else 'us-east-1')
                    end = datetime.now(timezone.utc)
                    start = end - timedelta(days=2)

                    # Object count
                    cnt_resp = cw.get_metric_statistics(
                        Namespace='AWS/S3',
                        MetricName='NumberOfObjects',
                        Dimensions=[
                            {'Name': 'BucketName', 'Value': name},
                            {'Name': 'StorageType', 'Value': 'AllStorageTypes'}
                        ],
                        StartTime=start, EndTime=end,
                        Period=86400, Statistics=['Average']
                    )
                    if cnt_resp['Datapoints']:
                        obj_count = int(sorted(cnt_resp['Datapoints'],
                            key=lambda x: x['Timestamp'])[-1]['Average'])

                    # Bucket size
                    size_resp = cw.get_metric_statistics(
                        Namespace='AWS/S3',
                        MetricName='BucketSizeBytes',
                        Dimensions=[
                            {'Name': 'BucketName', 'Value': name},
                            {'Name': 'StorageType', 'Value': 'StandardStorage'}
                        ],
                        StartTime=start, EndTime=end,
                        Period=86400, Statistics=['Average']
                    )
                    if size_resp['Datapoints']:
                        raw_size = sorted(size_resp['Datapoints'],
                            key=lambda x: x['Timestamp'])[-1]['Average']
                        total_size = fmt_bytes(raw_size)
                except:
                    pass

                # Try to get last modified from listing
                try:
                    s3r = self.session.resource('s3')
                    bucket = s3r.Bucket(name)
                    latest = None
                    for obj in bucket.objects.limit(1000):
                        if latest is None or obj.last_modified > latest:
                            latest = obj.last_modified
                    if latest:
                        last_modified = fmt_date(latest)
                except:
                    pass

                # Versioning
                versioning = safe(lambda: s3.get_bucket_versioning(
                    Bucket=name).get('Status', 'Disabled'), 'Disabled')

                # Public access
                public = 'Blocked'
                try:
                    pub = s3.get_public_access_block(Bucket=name)['PublicAccessBlockConfiguration']
                    if not all(pub.values()):
                        public = '⚠️ Partially Public'
                except:
                    public = '⚠️ Check manually'

                # Lifecycle
                has_lifecycle = 'No'
                try:
                    s3.get_bucket_lifecycle_configuration(Bucket=name)
                    has_lifecycle = 'Yes'
                except:
                    pass

                # Encryption
                encrypted = 'No'
                try:
                    s3.get_bucket_encryption(Bucket=name)
                    encrypted = 'Yes'
                except:
                    pass

                d = days_ago(created)
                empty = obj_count == 0 or obj_count == 'N/A'
                old = d and d > 180

                if empty and old:
                    rec = '⚠️ DELETE — Empty bucket older than 180 days'
                elif empty:
                    rec = '🔍 REVIEW — Empty bucket'
                else:
                    rec = '✅ KEEP — Has objects'

                rows.append({
                    'Resource Type': 'S3 Bucket',
                    'Name': name,
                    'Resource ID': name,
                    'Region': region,
                    'Status/State': public,
                    'Has Cost?': 'YES' if obj_count and obj_count != 'N/A' and obj_count > 0 else 'NO (empty)',
                    'Instance Type': 'N/A',
                    'Created / Launched': fmt_date(created),
                    'Last Used': last_modified,
                    'Days Old': d,
                    'Key Details': f"Objects: {obj_count} | Size: {total_size} | Versioning: {versioning} | Encrypted: {encrypted} | Lifecycle: {has_lifecycle}",
                    'Recommendation': rec,
                    'Notes': f"Public Access: {public}"
                })
        except ClientError as e:
            print(f"   ⚠️ S3 error: {e}")
        return rows

    # ── RDS ──────────────────────────────────
    def get_rds(self, region):
        rows = []
        try:
            rds = self.client('rds', region)
            dbs = rds.describe_db_instances()['DBInstances']
            for db in dbs:
                name = db['DBInstanceIdentifier']
                state = db.get('DBInstanceStatus', '')
                created = db.get('InstanceCreateTime')
                d = days_ago(created)
                multi_az = db.get('MultiAZ', False)
                storage_gb = db.get('AllocatedStorage', 0)

                if state == 'stopped':
                    rec = '🔍 REVIEW — RDS stopped (still costs for storage)'
                elif state == 'available':
                    rec = '✅ KEEP — Running'
                else:
                    rec = f'🔍 REVIEW — State: {state}'

                rows.append({
                    'Resource Type': 'RDS Instance',
                    'Name': name,
                    'Resource ID': db.get('DBInstanceArn', name),
                    'Region': region,
                    'Status/State': state,
                    'Has Cost?': 'YES',
                    'Instance Type': db.get('DBInstanceClass', 'N/A'),
                    'Created / Launched': fmt_date(created),
                    'Last Used': 'N/A (check CloudWatch)',
                    'Days Old': d,
                    'Key Details': f"Engine: {db.get('Engine')} {db.get('EngineVersion')} | Storage: {storage_gb}GB | Multi-AZ: {multi_az} | Endpoint: {db.get('Endpoint',{}).get('Address','N/A')}",
                    'Recommendation': rec,
                    'Notes': f"Backup: {db.get('BackupRetentionPeriod',0)} days"
                })
        except ClientError:
            pass
        return rows

    # ── LAMBDA ───────────────────────────────
    def get_lambda(self, region):
        rows = []
        try:
            lmb = self.client('lambda', region)
            paginator = lmb.get_paginator('list_functions')
            for page in paginator.paginate():
                for fn in page['Functions']:
                    name = fn['FunctionName']
                    modified = fn.get('LastModified', '')
                    modified_dt = None
                    try:
                        modified_dt = datetime.fromisoformat(modified.replace('Z', '+00:00'))
                    except:
                        pass
                    d_modified = days_ago(modified_dt)

                    # Check invocations via CloudWatch
                    invocations = 0
                    try:
                        cw = self.client('cloudwatch', region)
                        end = datetime.now(timezone.utc)
                        start = end - timedelta(days=30)
                        resp = cw.get_metric_statistics(
                            Namespace='AWS/Lambda',
                            MetricName='Invocations',
                            Dimensions=[{'Name': 'FunctionName', 'Value': name}],
                            StartTime=start, EndTime=end,
                            Period=2592000, Statistics=['Sum']
                        )
                        if resp['Datapoints']:
                            invocations = int(resp['Datapoints'][0]['Sum'])
                    except:
                        pass

                    if invocations == 0 and d_modified and d_modified > 90:
                        rec = '⚠️ DELETE — No invocations in 30 days, not modified in 90+ days'
                    elif invocations == 0:
                        rec = '🔍 REVIEW — No recent invocations'
                    else:
                        rec = f'✅ KEEP — {invocations} invocations (30d)'

                    rows.append({
                        'Resource Type': 'Lambda Function',
                        'Name': name,
                        'Resource ID': fn.get('FunctionArn', name),
                        'Region': region,
                        'Status/State': fn.get('State', 'Active'),
                        'Has Cost?': 'YES (per invocation)',
                        'Instance Type': f"{fn.get('Runtime','N/A')} | {fn.get('MemorySize',0)}MB",
                        'Created / Launched': 'N/A',
                        'Last Used': fmt_date(modified_dt),
                        'Days Old': d_modified,
                        'Key Details': f"Runtime: {fn.get('Runtime')} | Memory: {fn.get('MemorySize')}MB | Timeout: {fn.get('Timeout')}s | Invocations(30d): {invocations}",
                        'Recommendation': rec,
                        'Notes': fn.get('Handler', 'N/A')
                    })
        except ClientError:
            pass
        return rows

    # ── VPCs ─────────────────────────────────
    def get_vpcs(self, region):
        rows = []
        try:
            ec2 = self.client('ec2', region)
            vpcs = ec2.describe_vpcs()['Vpcs']
            for v in vpcs:
                name = get_name_tag(v.get('Tags')) or v['VpcId']
                is_default = v.get('IsDefault', False)
                # Count subnets
                subnets = ec2.describe_subnets(Filters=[
                    {'Name': 'vpc-id', 'Values': [v['VpcId']]}])['Subnets']
                # Count instances
                instances = ec2.describe_instances(Filters=[
                    {'Name': 'vpc-id', 'Values': [v['VpcId']]}])['Reservations']
                inst_count = sum(len(r['Instances']) for r in instances)

                if is_default and inst_count == 0:
                    rec = '🔍 REVIEW — Default VPC with no instances (consider cleaning)'
                elif inst_count == 0:
                    rec = '⚠️ DELETE — Empty custom VPC with no resources'
                else:
                    rec = f'✅ KEEP — {inst_count} instances'

                rows.append({
                    'Resource Type': 'VPC',
                    'Name': name,
                    'Resource ID': v['VpcId'],
                    'Region': region,
                    'Status/State': v.get('State', 'available'),
                    'Has Cost?': 'NO (VPC itself is free)',
                    'Instance Type': 'N/A',
                    'Created / Launched': 'N/A',
                    'Last Used': 'N/A',
                    'Days Old': 'N/A',
                    'Key Details': f"CIDR: {v.get('CidrBlock')} | Subnets: {len(subnets)} | Instances: {inst_count} | Default: {is_default}",
                    'Recommendation': rec,
                    'Notes': f"Tenancy: {v.get('InstanceTenancy','default')}"
                })
        except ClientError:
            pass
        return rows

    # ── SECURITY GROUPS ──────────────────────
    def get_security_groups(self, region):
        rows = []
        try:
            ec2 = self.client('ec2', region)
            sgs = ec2.describe_security_groups()['SecurityGroups']
            # Get all SGs in use
            used_sgs = set()
            try:
                for r in ec2.describe_instances()['Reservations']:
                    for inst in r['Instances']:
                        for sg in inst.get('SecurityGroups', []):
                            used_sgs.add(sg['GroupId'])
            except:
                pass
            try:
                for iface in ec2.describe_network_interfaces()['NetworkInterfaces']:
                    for sg in iface.get('Groups', []):
                        used_sgs.add(sg['GroupId'])
            except:
                pass

            for sg in sgs:
                sgid = sg['GroupId']
                name = sg.get('GroupName', sgid)
                in_use = sgid in used_sgs
                is_default = name == 'default'

                if is_default:
                    rec = '🔍 REVIEW — Default SG (cannot delete)'
                elif not in_use:
                    rec = '⚠️ DELETE — Not attached to any resource'
                else:
                    rec = '✅ KEEP — In use'

                inbound = len(sg.get('IpPermissions', []))
                outbound = len(sg.get('IpPermissionsEgress', []))

                rows.append({
                    'Resource Type': 'Security Group',
                    'Name': name,
                    'Resource ID': sgid,
                    'Region': region,
                    'Status/State': 'In Use' if in_use else 'Unused',
                    'Has Cost?': 'NO',
                    'Instance Type': 'N/A',
                    'Created / Launched': 'N/A',
                    'Last Used': 'In Use' if in_use else 'Never/Unknown',
                    'Days Old': 'N/A',
                    'Key Details': f"VPC: {sg.get('VpcId','N/A')} | Inbound rules: {inbound} | Outbound rules: {outbound}",
                    'Recommendation': rec,
                    'Notes': sg.get('Description', '')[:80]
                })
        except ClientError:
            pass
        return rows

    # ── ELASTIC IPs ──────────────────────────
    def get_eips(self, region):
        rows = []
        try:
            ec2 = self.client('ec2', region)
            eips = ec2.describe_addresses()['Addresses']
            for eip in eips:
                associated = bool(eip.get('AssociationId'))
                ip = eip.get('PublicIp', 'N/A')
                if not associated:
                    rec = '⚠️ DELETE — Unattached EIP costs $0.005/hr (~$3.60/month)'
                else:
                    rec = '✅ KEEP — Attached to resource'

                rows.append({
                    'Resource Type': 'Elastic IP',
                    'Name': ip,
                    'Resource ID': eip.get('AllocationId', ip),
                    'Region': region,
                    'Status/State': 'Attached' if associated else '⚠️ Unattached',
                    'Has Cost?': 'YES if unattached ($0.005/hr)',
                    'Instance Type': 'N/A',
                    'Created / Launched': 'N/A',
                    'Last Used': 'N/A',
                    'Days Old': 'N/A',
                    'Key Details': f"Public IP: {ip} | Attached to: {eip.get('InstanceId') or eip.get('NetworkInterfaceId','Not attached')}",
                    'Recommendation': rec,
                    'Notes': eip.get('Domain', 'vpc')
                })
        except ClientError:
            pass
        return rows

    # ── NAT GATEWAYS ─────────────────────────
    def get_nat_gateways(self, region):
        rows = []
        try:
            ec2 = self.client('ec2', region)
            nats = ec2.describe_nat_gateways()['NatGateways']
            for nat in nats:
                if nat.get('State') == 'deleted':
                    continue
                name = get_name_tag(nat.get('Tags')) or nat['NatGatewayId']
                state = nat.get('State', '')
                created = nat.get('CreateTime')
                d = days_ago(created)

                # Check traffic via CloudWatch
                bytes_out = 0
                try:
                    cw = self.client('cloudwatch', region)
                    end = datetime.now(timezone.utc)
                    start = end - timedelta(days=7)
                    resp = cw.get_metric_statistics(
                        Namespace='AWS/NATGateway',
                        MetricName='BytesOutToDestination',
                        Dimensions=[{'Name': 'NatGatewayId', 'Value': nat['NatGatewayId']}],
                        StartTime=start, EndTime=end,
                        Period=604800, Statistics=['Sum']
                    )
                    if resp['Datapoints']:
                        bytes_out = int(resp['Datapoints'][0]['Sum'])
                except:
                    pass

                if state == 'available' and bytes_out == 0:
                    rec = '⚠️ DELETE — No traffic in 7 days ($0.045/hr + data charges)'
                elif state == 'available':
                    rec = '✅ KEEP — Active traffic'
                else:
                    rec = f'🔍 REVIEW — State: {state}'

                rows.append({
                    'Resource Type': 'NAT Gateway',
                    'Name': name,
                    'Resource ID': nat['NatGatewayId'],
                    'Region': region,
                    'Status/State': state,
                    'Has Cost?': 'YES ($0.045/hr + $0.045/GB)',
                    'Instance Type': 'N/A',
                    'Created / Launched': fmt_date(created),
                    'Last Used': f"Traffic(7d): {fmt_bytes(bytes_out)}",
                    'Days Old': d,
                    'Key Details': f"Subnet: {nat.get('SubnetId','N/A')} | VPC: {nat.get('VpcId','N/A')} | Traffic 7d: {fmt_bytes(bytes_out)}",
                    'Recommendation': rec,
                    'Notes': ''
                })
        except ClientError:
            pass
        return rows

    # ── LOAD BALANCERS ───────────────────────
    def get_elb(self, region):
        rows = []
        try:
            elb = self.client('elbv2', region)
            lbs = elb.describe_load_balancers()['LoadBalancers']
            for lb in lbs:
                name = lb['LoadBalancerName']
                state = lb.get('State', {}).get('Code', '')
                created = lb.get('CreatedTime')
                d = days_ago(created)

                # Check request count
                requests = 0
                try:
                    cw = self.client('cloudwatch', region)
                    end = datetime.now(timezone.utc)
                    start = end - timedelta(days=7)
                    metric = 'RequestCount' if lb.get('Type') == 'application' else 'ActiveFlowCount'
                    lb_dim = lb['LoadBalancerArn'].split(':loadbalancer/')[-1]
                    resp = cw.get_metric_statistics(
                        Namespace='AWS/ApplicationELB' if lb.get('Type') == 'application' else 'AWS/NetworkELB',
                        MetricName=metric,
                        Dimensions=[{'Name': 'LoadBalancer', 'Value': lb_dim}],
                        StartTime=start, EndTime=end,
                        Period=604800, Statistics=['Sum']
                    )
                    if resp['Datapoints']:
                        requests = int(resp['Datapoints'][0]['Sum'])
                except:
                    pass

                if state == 'active' and requests == 0:
                    rec = '⚠️ DELETE — No traffic in 7 days (~$16-22/month minimum)'
                elif state == 'active':
                    rec = f'✅ KEEP — {requests:,} requests (7d)'
                else:
                    rec = f'🔍 REVIEW — State: {state}'

                rows.append({
                    'Resource Type': f"Load Balancer ({lb.get('Type','').upper()})",
                    'Name': name,
                    'Resource ID': lb['LoadBalancerArn'],
                    'Region': region,
                    'Status/State': state,
                    'Has Cost?': 'YES (~$0.008/hr + LCU charges)',
                    'Instance Type': lb.get('Type', 'N/A'),
                    'Created / Launched': fmt_date(created),
                    'Last Used': f"Requests(7d): {requests:,}",
                    'Days Old': d,
                    'Key Details': f"DNS: {lb.get('DNSName','N/A')} | Scheme: {lb.get('Scheme','N/A')} | Requests 7d: {requests:,}",
                    'Recommendation': rec,
                    'Notes': ''
                })
        except ClientError:
            pass
        return rows

    # ── EBS VOLUMES ──────────────────────────
    def get_ebs(self, region):
        rows = []
        try:
            ec2 = self.client('ec2', region)
            volumes = ec2.describe_volumes()['Volumes']
            for vol in volumes:
                name = get_name_tag(vol.get('Tags')) or vol['VolumeId']
                state = vol.get('State', '')
                created = vol.get('CreateTime')
                d = days_ago(created)
                size_gb = vol.get('Size', 0)
                attached = len(vol.get('Attachments', [])) > 0
                vol_type = vol.get('VolumeType', 'gp2')

                cost_per_gb = {'gp2': 0.10, 'gp3': 0.08, 'io1': 0.125,
                               'io2': 0.125, 'st1': 0.045, 'sc1': 0.025}.get(vol_type, 0.10)
                monthly_cost = size_gb * cost_per_gb

                if state == 'available' and not attached:
                    rec = f'⚠️ DELETE — Unattached volume costing ~${monthly_cost:.2f}/month'
                elif attached:
                    rec = '✅ KEEP — Attached to instance'
                else:
                    rec = '🔍 REVIEW — Check state'

                rows.append({
                    'Resource Type': 'EBS Volume',
                    'Name': name,
                    'Resource ID': vol['VolumeId'],
                    'Region': region,
                    'Status/State': state,
                    'Has Cost?': f'YES (~${monthly_cost:.2f}/month)',
                    'Instance Type': f"{vol_type} | {size_gb}GB",
                    'Created / Launched': fmt_date(created),
                    'Last Used': vol.get('Attachments', [{}])[0].get('AttachTime', 'Never attached') if vol.get('Attachments') else 'Never attached',
                    'Days Old': d,
                    'Key Details': f"Size: {size_gb}GB | Type: {vol_type} | IOPS: {vol.get('Iops','N/A')} | Attached: {attached} | AZ: {vol.get('AvailabilityZone','N/A')}",
                    'Recommendation': rec,
                    'Notes': f"Est. cost: ${monthly_cost:.2f}/month"
                })
        except ClientError:
            pass
        return rows

    # ── EBS SNAPSHOTS ─────────────────────────
    def get_snapshots(self, region):
        rows = []
        try:
            ec2 = self.client('ec2', region)
            snapshots = ec2.describe_snapshots(OwnerIds=['self'])['Snapshots']
            for snap in snapshots:
                name = get_name_tag(snap.get('Tags')) or snap['SnapshotId']
                created = snap.get('StartTime')
                d = days_ago(created)
                size_gb = snap.get('VolumeSize', 0)
                monthly_cost = size_gb * 0.05

                if d and d > 365:
                    rec = f'⚠️ DELETE — Snapshot older than 1 year (~${monthly_cost:.2f}/month)'
                elif d and d > 90:
                    rec = f'🔍 REVIEW — Snapshot {d} days old'
                else:
                    rec = '✅ KEEP — Recent snapshot'

                rows.append({
                    'Resource Type': 'EBS Snapshot',
                    'Name': name,
                    'Resource ID': snap['SnapshotId'],
                    'Region': region,
                    'Status/State': snap.get('State', ''),
                    'Has Cost?': f'YES ($0.05/GB/month = ~${monthly_cost:.2f}/month)',
                    'Instance Type': f"{size_gb}GB",
                    'Created / Launched': fmt_date(created),
                    'Last Used': fmt_date(created),
                    'Days Old': d,
                    'Key Details': f"Size: {size_gb}GB | Volume: {snap.get('VolumeId','N/A')} | Encrypted: {snap.get('Encrypted',False)}",
                    'Recommendation': rec,
                    'Notes': snap.get('Description', '')[:80]
                })
        except ClientError:
            pass
        return rows

    # ── IAM USERS ────────────────────────────
    def get_iam_users(self):
        rows = []
        try:
            iam = self.client('iam')
            paginator = iam.get_paginator('list_users')
            for page in paginator.paginate():
                for user in page['Users']:
                    name = user['UserName']
                    created = user.get('CreateDate')
                    d = days_ago(created)
                    last_used = user.get('PasswordLastUsed')
                    d_last = days_ago(last_used)

                    # Access keys
                    keys = iam.list_access_keys(UserName=name)['AccessKeyMetadata']
                    active_keys = [k for k in keys if k['Status'] == 'Active']
                    old_keys = [k for k in active_keys if days_ago(k.get('CreateDate')) and days_ago(k.get('CreateDate')) > 90]

                    if d_last and d_last > 90:
                        rec = f'⚠️ DELETE — No login in {d_last} days'
                    elif old_keys:
                        rec = f'🔍 REVIEW — Access key older than 90 days'
                    elif not last_used:
                        rec = '🔍 REVIEW — Never logged in'
                    else:
                        rec = '✅ KEEP — Active user'

                    rows.append({
                        'Resource Type': 'IAM User',
                        'Name': name,
                        'Resource ID': user.get('UserId', name),
                        'Region': 'Global',
                        'Status/State': 'Active' if last_used else 'Never logged in',
                        'Has Cost?': 'NO',
                        'Instance Type': 'N/A',
                        'Created / Launched': fmt_date(created),
                        'Last Used': fmt_date(last_used) if last_used else 'Never',
                        'Days Old': d,
                        'Key Details': f"Active Keys: {len(active_keys)} | Old Keys (>90d): {len(old_keys)} | MFA: check manually",
                        'Recommendation': rec,
                        'Notes': f"ARN: {user.get('Arn','N/A')}"
                    })
        except ClientError:
            pass
        return rows

    # ── RDS SNAPSHOTS ─────────────────────────
    def get_rds_snapshots(self, region):
        rows = []
        try:
            rds = self.client('rds', region)
            snaps = rds.describe_db_snapshots(SnapshotType='manual')['DBSnapshots']
            for snap in snaps:
                name = snap.get('DBSnapshotIdentifier', '')
                created = snap.get('SnapshotCreateTime')
                d = days_ago(created)
                size_gb = snap.get('AllocatedStorage', 0)
                monthly_cost = size_gb * 0.095

                if d and d > 365:
                    rec = f'⚠️ DELETE — Manual snapshot older than 1 year'
                elif d and d > 90:
                    rec = f'🔍 REVIEW — {d} days old'
                else:
                    rec = '✅ KEEP — Recent snapshot'

                rows.append({
                    'Resource Type': 'RDS Snapshot',
                    'Name': name,
                    'Resource ID': snap.get('DBSnapshotArn', name),
                    'Region': region,
                    'Status/State': snap.get('Status', ''),
                    'Has Cost?': f'YES (~${monthly_cost:.2f}/month)',
                    'Instance Type': f"{snap.get('Engine','')} | {size_gb}GB",
                    'Created / Launched': fmt_date(created),
                    'Last Used': fmt_date(created),
                    'Days Old': d,
                    'Key Details': f"DB: {snap.get('DBInstanceIdentifier','N/A')} | Engine: {snap.get('Engine','N/A')} | Size: {size_gb}GB | Encrypted: {snap.get('Encrypted',False)}",
                    'Recommendation': rec,
                    'Notes': ''
                })
        except ClientError:
            pass
        return rows

    # ── VPN CONNECTIONS ──────────────────────
    def get_vpn(self, region):
        rows = []
        try:
            ec2 = self.client('ec2', region)
            vpns = ec2.describe_vpn_connections()['VpnConnections']
            for vpn in vpns:
                if vpn.get('State') == 'deleted':
                    continue
                name = get_name_tag(vpn.get('Tags')) or vpn['VpnConnectionId']
                state = vpn.get('State', '')
                tunnels = vpn.get('VgwTelemetry', [])
                up_tunnels = [t for t in tunnels if t.get('Status') == 'UP']

                rows.append({
                    'Resource Type': 'VPN Connection',
                    'Name': name,
                    'Resource ID': vpn['VpnConnectionId'],
                    'Region': region,
                    'Status/State': state,
                    'Has Cost?': 'YES ($0.05/hr = ~$36/month)',
                    'Instance Type': vpn.get('Type', 'ipsec.1'),
                    'Created / Launched': 'N/A',
                    'Last Used': 'N/A',
                    'Days Old': 'N/A',
                    'Key Details': f"CGW: {vpn.get('CustomerGatewayId','N/A')} | VGW: {vpn.get('VpnGatewayId','N/A')} | Tunnels UP: {len(up_tunnels)}/{len(tunnels)}",
                    'Recommendation': '✅ KEEP — Active VPN' if state == 'available' else f'🔍 REVIEW — State: {state}',
                    'Notes': f"Tunnel IPs: {', '.join([t.get('OutsideIpAddress','') for t in tunnels])}"
                })
        except ClientError:
            pass
        return rows

    # ── ELASTICACHE ──────────────────────────
    def get_elasticache(self, region):
        rows = []
        try:
            ec = self.client('elasticache', region)
            clusters = ec.describe_cache_clusters()['CacheClusters']
            for c in clusters:
                cid = c['CacheClusterId']
                state = c.get('CacheClusterStatus', '')
                created = c.get('CacheClusterCreateTime')
                d = days_ago(created)
                node_type = c.get('CacheNodeType', '')
                rows.append({
                    'Resource Type': 'ElastiCache Cluster',
                    'Name': cid,
                    'Resource ID': cid,
                    'Region': region,
                    'Status/State': state,
                    'Has Cost?': 'YES (per node/hour)',
                    'Instance Type': node_type,
                    'Created / Launched': fmt_date(created),
                    'Last Used': 'N/A (check CloudWatch)',
                    'Days Old': d,
                    'Key Details': f"Engine: {c.get('Engine')} {c.get('EngineVersion')} | Nodes: {c.get('NumCacheNodes')} | Node type: {node_type}",
                    'Recommendation': '✅ KEEP — Active' if state == 'available' else f'🔍 REVIEW — State: {state}',
                    'Notes': ''
                })
        except ClientError:
            pass
        return rows

    # ── MAIN SCAN ─────────────────────────────
    def scan_all(self):
        print('\n' + '='*60)
        print('  🔍 Getting AWS account info...')

        try:
            sts = self.client('sts')
            identity = sts.get_caller_identity()
            self.account_id = identity.get('Account', 'unknown')
            print(f'  ✅ Account: {self.account_id}')
        except Exception as e:
            print(f'  ❌ Auth failed: {e}')
            sys.exit(1)

        print('\n  🌍 Finding active regions...')
        all_regions = self.get_regions()
        active_regions = []
        for region in all_regions:
            sys.stdout.write(f'     Checking {region}... ')
            sys.stdout.flush()
            if self.is_active_region(region):
                print('✅')
                active_regions.append(region)
            else:
                print('⬜ skip')

        all_rows = []

        print(f'\n  📊 Scanning {len(active_regions)} active regions...\n')

        def scan_region(region):
            r = []
            r += self.get_ec2(region)
            r += self.get_vpcs(region)
            r += self.get_security_groups(region)
            r += self.get_eips(region)
            r += self.get_nat_gateways(region)
            r += self.get_elb(region)
            r += self.get_ebs(region)
            r += self.get_snapshots(region)
            r += self.get_rds(region)
            r += self.get_rds_snapshots(region)
            r += self.get_lambda(region)
            r += self.get_elasticache(region)
            r += self.get_vpn(region)
            print(f'     ✅ {region}: {len(r)} resources found')
            return r

        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = {executor.submit(scan_region, r): r for r in active_regions}
            for future in as_completed(futures):
                try:
                    all_rows.extend(future.result())
                except Exception as e:
                    print(f'     ⚠️ Error: {e}')

        # Global resources (S3, IAM)
        print('\n  🪣 Scanning S3 buckets (global)...')
        all_rows += self.get_s3()

        print('\n  👤 Scanning IAM users (global)...')
        all_rows += self.get_iam_users()

        return all_rows, self.account_id, active_regions


# ─────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────
def safe_val(v):
    """Strip timezone from datetime objects before writing to Excel."""
    if v is None:
        return 'N/A'
    if hasattr(v, 'tzinfo') and v.tzinfo is not None:
        return v.replace(tzinfo=None).strftime('%Y-%m-%d')
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return v


def build_excel(rows, account_id, regions, company, output_path):
    wb = Workbook()

    # Sanitize all values — strip timezone from any datetime objects
    clean_rows = []
    for row in rows:
        clean = {}
        for k, v in row.items():
            if hasattr(v, 'tzinfo') and v.tzinfo is not None:
                clean[k] = v.replace(tzinfo=None).strftime('%Y-%m-%d')
            elif isinstance(v, datetime):
                clean[k] = v.strftime('%Y-%m-%d')
            else:
                clean[k] = v
        clean_rows.append(clean)
    rows = clean_rows

    scanned_at = datetime.now().strftime('%Y-%m-%d %H:%M UTC')

    # ── SHEET 1: SUMMARY DASHBOARD ────────────
    ws_sum = wb.active
    ws_sum.title = '📊 Summary'
    style_title(ws_sum, f'{company} — AWS Resource Report',
                f'Account: {account_id}  |  Scanned: {scanned_at}  |  Regions: {", ".join(regions[:5])}{"..." if len(regions)>5 else ""}')

    # Summary stats
    ws_sum.row_dimensions[4].height = 18
    ws_sum['A4'] = 'SUMMARY STATS'
    ws_sum['A4'].font = Font(name='Arial', bold=True, size=11, color='1F3864')

    summary_headers = ['Metric', 'Count']
    for i, h in enumerate(summary_headers, 1):
        c = ws_sum.cell(row=5, column=i, value=h)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', start_color='1F3864')
        c.alignment = Alignment(horizontal='center')

    # Count by type
    type_counts = {}
    cost_yes = 0
    delete_rec = 0
    for row in rows:
        t = row['Resource Type']
        type_counts[t] = type_counts.get(t, 0) + 1
        if 'YES' in str(row.get('Has Cost?', '')):
            cost_yes += 1
        if 'DELETE' in str(row.get('Recommendation', '')):
            delete_rec += 1

    summary_data = [
        ('Total Resources Found', len(rows)),
        ('Resources with Cost', cost_yes),
        ('Recommended to Delete', delete_rec),
        ('Active Regions Scanned', len(regions)),
        ('Report Generated', scanned_at),
    ]
    for i, (k, v) in enumerate(summary_data):
        row_n = 6 + i
        ws_sum.cell(row=row_n, column=1, value=k).font = Font(name='Arial', size=10, bold=True)
        c = ws_sum.cell(row=row_n, column=2, value=v)
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(horizontal='center')
        if i == 2 and isinstance(v, int) and v > 0:
            c.font = Font(name='Arial', size=10, bold=True, color='C00000')

    # Type breakdown table
    ws_sum.cell(row=5, column=4, value='Resource Type').font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ws_sum.cell(row=5, column=4).fill = PatternFill('solid', start_color='1F3864')
    ws_sum.cell(row=5, column=4).alignment = Alignment(horizontal='center')
    ws_sum.cell(row=5, column=5, value='Count').font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ws_sum.cell(row=5, column=5).fill = PatternFill('solid', start_color='1F3864')
    ws_sum.cell(row=5, column=5).alignment = Alignment(horizontal='center')

    for i, (t, cnt) in enumerate(sorted(type_counts.items(), key=lambda x: -x[1])):
        row_n = 6 + i
        ws_sum.cell(row=row_n, column=4, value=t).font = Font(name='Arial', size=9)
        c = ws_sum.cell(row=row_n, column=5, value=cnt)
        c.font = Font(name='Arial', size=9, bold=True)
        c.alignment = Alignment(horizontal='center')

    # Column widths
    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 20
    ws_sum.column_dimensions['C'].width = 5
    ws_sum.column_dimensions['D'].width = 30
    ws_sum.column_dimensions['E'].width = 10

    # ── SHEET 2: ALL RESOURCES ────────────────
    ws_all = wb.create_sheet('📋 All Resources')
    style_title(ws_all, f'All AWS Resources — {company}',
                f'Account: {account_id}  |  Total: {len(rows)} resources  |  Scanned: {scanned_at}')

    headers = [
        'Resource Type', 'Name', 'Resource ID', 'Region',
        'Status / State', 'Has Cost?', 'Instance/Size Type',
        'Created / Launched', 'Last Used', 'Days Old',
        'Key Details', 'Recommendation', 'Notes'
    ]

    hdr_row = 3
    for col, h in enumerate(headers, 1):
        style_header(ws_all, hdr_row, len(headers))
        ws_all.cell(row=hdr_row, column=col).value = h
    ws_all.row_dimensions[hdr_row].height = 30

    for row_idx, row in enumerate(rows, hdr_row + 1):
        even = (row_idx % 2 == 0)
        style_row(ws_all, row_idx, len(headers), even)
        ws_all.cell(row=row_idx, column=1).value = row.get('Resource Type', '')
        ws_all.cell(row=row_idx, column=2).value = row.get('Name', '')
        ws_all.cell(row=row_idx, column=3).value = row.get('Resource ID', '')
        ws_all.cell(row=row_idx, column=4).value = row.get('Region', '')
        ws_all.cell(row=row_idx, column=5).value = row.get('Status/State', '')
        cost_cell = ws_all.cell(row=row_idx, column=6, value=row.get('Has Cost?', ''))
        cost_color(ws_all, cost_cell, 'YES' if 'YES' in str(row.get('Has Cost?', '')) else 'NO')
        ws_all.cell(row=row_idx, column=7).value = row.get('Instance Type', '')
        ws_all.cell(row=row_idx, column=8).value = row.get('Created / Launched', '')
        ws_all.cell(row=row_idx, column=9).value = row.get('Last Used', '')
        ws_all.cell(row=row_idx, column=10).value = row.get('Days Old', '')
        ws_all.cell(row=row_idx, column=11).value = row.get('Key Details', '')
        rec_cell = ws_all.cell(row=row_idx, column=12, value=row.get('Recommendation', ''))
        recommend_color(rec_cell, row.get('Recommendation', ''))
        ws_all.cell(row=row_idx, column=13).value = row.get('Notes', '')

    # Column widths
    col_widths = [22, 28, 32, 14, 14, 22, 22, 16, 20, 10, 55, 42, 30]
    for i, w in enumerate(col_widths, 1):
        ws_all.column_dimensions[get_column_letter(i)].width = w

    ws_all.freeze_panes = 'A4'
    ws_all.auto_filter.ref = f'A{hdr_row}:{get_column_letter(len(headers))}{hdr_row}'

    # ── SHEET 3: DELETE RECOMMENDATIONS ───────
    delete_rows = [r for r in rows if 'DELETE' in str(r.get('Recommendation', ''))]
    ws_del = wb.create_sheet('🗑️ Delete Recommendations')
    style_title(ws_del, f'⚠️ Resources Recommended for Deletion — {company}',
                f'{len(delete_rows)} resources found  |  Account: {account_id}  |  Scanned: {scanned_at}')

    ws_del.cell(row=hdr_row, column=1)
    for col, h in enumerate(headers, 1):
        style_header(ws_del, hdr_row, len(headers))
        ws_del.cell(row=hdr_row, column=col).value = h
    ws_del.row_dimensions[hdr_row].height = 30

    for row_idx, row in enumerate(delete_rows, hdr_row + 1):
        even = (row_idx % 2 == 0)
        style_row(ws_del, row_idx, len(headers), even)
        ws_del.cell(row=row_idx, column=1).value = row.get('Resource Type', '')
        ws_del.cell(row=row_idx, column=2).value = row.get('Name', '')
        ws_del.cell(row=row_idx, column=3).value = row.get('Resource ID', '')
        ws_del.cell(row=row_idx, column=4).value = row.get('Region', '')
        ws_del.cell(row=row_idx, column=5).value = row.get('Status/State', '')
        cc = ws_del.cell(row=row_idx, column=6, value=row.get('Has Cost?', ''))
        cost_color(ws_del, cc, 'YES' if 'YES' in str(row.get('Has Cost?', '')) else 'NO')
        ws_del.cell(row=row_idx, column=7).value = row.get('Instance Type', '')
        ws_del.cell(row=row_idx, column=8).value = row.get('Created / Launched', '')
        ws_del.cell(row=row_idx, column=9).value = row.get('Last Used', '')
        ws_del.cell(row=row_idx, column=10).value = row.get('Days Old', '')
        ws_del.cell(row=row_idx, column=11).value = row.get('Key Details', '')
        rc = ws_del.cell(row=row_idx, column=12, value=row.get('Recommendation', ''))
        recommend_color(rc, row.get('Recommendation', ''))
        ws_del.cell(row=row_idx, column=13).value = row.get('Notes', '')

    for i, w in enumerate(col_widths, 1):
        ws_del.column_dimensions[get_column_letter(i)].width = w
    ws_del.freeze_panes = 'A4'
    ws_del.auto_filter.ref = f'A{hdr_row}:{get_column_letter(len(headers))}{hdr_row}'

    # ── SHEET 4: S3 DETAILED ──────────────────
    s3_rows = [r for r in rows if r.get('Resource Type') == 'S3 Bucket']
    ws_s3 = wb.create_sheet('🪣 S3 Buckets Detail')
    style_title(ws_s3, f'S3 Buckets — {company}',
                f'{len(s3_rows)} buckets  |  Account: {account_id}  |  Scanned: {scanned_at}')

    s3_headers = ['Bucket Name', 'Region', 'Created', 'Last Modified',
                  'Object Count', 'Total Size', 'Versioning', 'Encrypted',
                  'Lifecycle', 'Public Access', 'Has Cost?', 'Recommendation']

    for col, h in enumerate(s3_headers, 1):
        style_header(ws_s3, hdr_row, len(s3_headers))
        ws_s3.cell(row=hdr_row, column=col).value = h
    ws_s3.row_dimensions[hdr_row].height = 30

    for row_idx, row in enumerate(s3_rows, hdr_row + 1):
        even = (row_idx % 2 == 0)
        style_row(ws_s3, row_idx, len(s3_headers), even)
        details = row.get('Key Details', '')
        # Parse details string
        obj_count = 'N/A'
        size = 'N/A'
        versioning = 'N/A'
        encrypted = 'N/A'
        lifecycle = 'N/A'

        for part in details.split('|'):
            part = part.strip()
            if part.startswith('Objects:'):
                obj_count = part.split(':')[1].strip()
            elif part.startswith('Size:'):
                size = part.split(':', 1)[1].strip()
            elif part.startswith('Versioning:'):
                versioning = part.split(':')[1].strip()
            elif part.startswith('Encrypted:'):
                encrypted = part.split(':')[1].strip()
            elif part.startswith('Lifecycle:'):
                lifecycle = part.split(':')[1].strip()

        ws_s3.cell(row=row_idx, column=1).value = row.get('Name', '')
        ws_s3.cell(row=row_idx, column=2).value = row.get('Region', '')
        ws_s3.cell(row=row_idx, column=3).value = row.get('Created / Launched', '')
        ws_s3.cell(row=row_idx, column=4).value = row.get('Last Used', '')
        ws_s3.cell(row=row_idx, column=5).value = obj_count
        ws_s3.cell(row=row_idx, column=6).value = size
        ws_s3.cell(row=row_idx, column=7).value = versioning
        enc_cell = ws_s3.cell(row=row_idx, column=8, value=encrypted)
        if encrypted == 'No':
            enc_cell.font = Font(name='Arial', bold=True, size=9, color='C00000')
        ws_s3.cell(row=row_idx, column=9).value = lifecycle
        pub_cell = ws_s3.cell(row=row_idx, column=10, value=row.get('Status/State', ''))
        if 'Public' in str(row.get('Status/State', '')):
            pub_cell.font = Font(name='Arial', bold=True, size=9, color='C00000')
            pub_cell.fill = PatternFill('solid', start_color='FFE7E7')
        cc = ws_s3.cell(row=row_idx, column=11, value=row.get('Has Cost?', ''))
        cost_color(ws_s3, cc, 'YES' if 'YES' in str(row.get('Has Cost?', '')) else 'NO')
        rc = ws_s3.cell(row=row_idx, column=12, value=row.get('Recommendation', ''))
        recommend_color(rc, row.get('Recommendation', ''))

    s3_widths = [40, 15, 14, 18, 14, 14, 14, 12, 12, 22, 22, 42]
    for i, w in enumerate(s3_widths, 1):
        ws_s3.column_dimensions[get_column_letter(i)].width = w
    ws_s3.freeze_panes = 'A4'
    ws_s3.auto_filter.ref = f'A{hdr_row}:{get_column_letter(len(s3_headers))}{hdr_row}'

    # ── SHEET 5: COST RESOURCES ───────────────
    cost_rows = [r for r in rows if 'YES' in str(r.get('Has Cost?', ''))]
    ws_cost = wb.create_sheet('💰 Costing Resources')
    style_title(ws_cost, f'Resources with Costs — {company}',
                f'{len(cost_rows)} costing resources  |  Account: {account_id}')

    for col, h in enumerate(headers, 1):
        style_header(ws_cost, hdr_row, len(headers))
        ws_cost.cell(row=hdr_row, column=col).value = h
    ws_cost.row_dimensions[hdr_row].height = 30

    for row_idx, row in enumerate(cost_rows, hdr_row + 1):
        even = (row_idx % 2 == 0)
        style_row(ws_cost, row_idx, len(headers), even)
        ws_cost.cell(row=row_idx, column=1).value = row.get('Resource Type', '')
        ws_cost.cell(row=row_idx, column=2).value = row.get('Name', '')
        ws_cost.cell(row=row_idx, column=3).value = row.get('Resource ID', '')
        ws_cost.cell(row=row_idx, column=4).value = row.get('Region', '')
        ws_cost.cell(row=row_idx, column=5).value = row.get('Status/State', '')
        cc = ws_cost.cell(row=row_idx, column=6, value=row.get('Has Cost?', ''))
        cost_color(ws_cost, cc, 'YES')
        ws_cost.cell(row=row_idx, column=7).value = row.get('Instance Type', '')
        ws_cost.cell(row=row_idx, column=8).value = row.get('Created / Launched', '')
        ws_cost.cell(row=row_idx, column=9).value = row.get('Last Used', '')
        ws_cost.cell(row=row_idx, column=10).value = row.get('Days Old', '')
        ws_cost.cell(row=row_idx, column=11).value = row.get('Key Details', '')
        rc = ws_cost.cell(row=row_idx, column=12, value=row.get('Recommendation', ''))
        recommend_color(rc, row.get('Recommendation', ''))
        ws_cost.cell(row=row_idx, column=13).value = row.get('Notes', '')

    for i, w in enumerate(col_widths, 1):
        ws_cost.column_dimensions[get_column_letter(i)].width = w
    ws_cost.freeze_panes = 'A4'
    ws_cost.auto_filter.ref = f'A{hdr_row}:{get_column_letter(len(headers))}{hdr_row}'

    # ── SHEET 6: FREE RESOURCES ───────────────
    free_rows = [r for r in rows if 'YES' not in str(r.get('Has Cost?', ''))]
    ws_free = wb.create_sheet('✅ Free Resources')
    style_title(ws_free, f'Free Resources (No Direct Cost) — {company}',
                f'{len(free_rows)} free resources  |  Account: {account_id}')

    for col, h in enumerate(headers, 1):
        style_header(ws_free, hdr_row, len(headers))
        ws_free.cell(row=hdr_row, column=col).value = h
    ws_free.row_dimensions[hdr_row].height = 30

    for row_idx, row in enumerate(free_rows, hdr_row + 1):
        even = (row_idx % 2 == 0)
        style_row(ws_free, row_idx, len(headers), even)
        ws_free.cell(row=row_idx, column=1).value = row.get('Resource Type', '')
        ws_free.cell(row=row_idx, column=2).value = row.get('Name', '')
        ws_free.cell(row=row_idx, column=3).value = row.get('Resource ID', '')
        ws_free.cell(row=row_idx, column=4).value = row.get('Region', '')
        ws_free.cell(row=row_idx, column=5).value = row.get('Status/State', '')
        cc = ws_free.cell(row=row_idx, column=6, value=row.get('Has Cost?', ''))
        cost_color(ws_free, cc, 'NO')
        ws_free.cell(row=row_idx, column=7).value = row.get('Instance Type', '')
        ws_free.cell(row=row_idx, column=8).value = row.get('Created / Launched', '')
        ws_free.cell(row=row_idx, column=9).value = row.get('Last Used', '')
        ws_free.cell(row=row_idx, column=10).value = row.get('Days Old', '')
        ws_free.cell(row=row_idx, column=11).value = row.get('Key Details', '')
        rc = ws_free.cell(row=row_idx, column=12, value=row.get('Recommendation', ''))
        recommend_color(rc, row.get('Recommendation', ''))
        ws_free.cell(row=row_idx, column=13).value = row.get('Notes', '')

    for i, w in enumerate(col_widths, 1):
        ws_free.column_dimensions[get_column_letter(i)].width = w
    ws_free.freeze_panes = 'A4'
    ws_free.auto_filter.ref = f'A{hdr_row}:{get_column_letter(len(headers))}{hdr_row}'

    wb.save(output_path)
    print(f'\n✅ Excel report saved: {output_path}')


# ─────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='AWS Resource Excel Reporter — inventory, cost, and cleanup recommendations'
    )
    parser.add_argument('--access-key', '-a', required=True)
    parser.add_argument('--secret-key', '-s', required=True)
    parser.add_argument('--session-token', '-t', default=None)
    parser.add_argument('--output', '-o', default='aws-resource-report')
    args = parser.parse_args()

    print('=' * 60)
    print('  AWS Resource Excel Reporter')
    print('=' * 60)

    company = input('\n  Enter your company name (or press Enter to skip): ').strip()
    if not company:
        company = 'AWS Account'

    reporter = AWSExcelReporter(args.access_key, args.secret_key, args.session_token)
    rows, account_id, regions = reporter.scan_all()

    output_path = f'{args.output}-{account_id}.xlsx'

    print(f'\n  📊 Building Excel report...')
    build_excel(rows, account_id, regions, company, output_path)

    print('\n' + '=' * 60)
    print(f'  ✅ Done!')
    print(f'  📊 Total Resources   : {len(rows)}')
    delete_count = sum(1 for r in rows if 'DELETE' in str(r.get('Recommendation', '')))
    cost_count = sum(1 for r in rows if 'YES' in str(r.get('Has Cost?', '')))
    print(f'  💰 Costing Resources : {cost_count}')
    print(f'  🗑️  Delete Recommended: {delete_count}')
    print(f'  🌍 Regions Scanned   : {len(regions)}')
    print(f'  📁 Output File       : {output_path}')
    print('=' * 60)
    print(f'\n  Open {output_path} in Excel or Google Sheets!\n')


if __name__ == '__main__':
    main()
